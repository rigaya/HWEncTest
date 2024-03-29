﻿# -*- coding: utf-8 -*-

import subprocess
import shlex
import os
import shutil
import sys
import ctypes
import openpyxl
import traceback
import datetime
import codecs
import difflib
import psutil
import time
import threading
import joblib

outputdir = ""
logpath = ""
input_xlsx = ""
output_xlsx_filename = "0000_result.xlsx"
encoder_log_path = ""
encoder_log_prefix = ".log"

x264_path = "x264"
avs2pipemod_path = ""
qsvencc_path = "qsvencc"
nvencc_path = "nvencc"
vceencc_path = "vceencc"
rkmppenc_path = "rkmppenc"
ffmpeg_path = "ffmpeg"
mediainfo_path = "mediainfo"

x264_path_win = r'x64\x264.exe'
avs2pipemod_path_win = r'x86\Avs2Pipemod.exe'
qsvencc_path_win = r'x64\QSVEncC64.exe'
nvencc_path_win = r'x64\NVEncC64.exe'
vceencc_path_win = r'x64\VCEEncC64.exe'
ffmpeg_path_win = r'x64\ffmpeg.exe'
mediainfo_path_win = r'MediaInfo\MediaInfo.exe'

mediainfo_template = r'HWEncTestMediaInfoTemplate.txt'
mediainfo_check = False
mediainfo_check_log_appendix = ".mediainfo.txt"
mediainfo_check_diff_appendix = ".mediainfo.diff"
encoder_path = qsvencc_path
encoder_name = ""
filesize_threshold = 0.25
test_count=0
test_start=0
test_target=0
UseProcessChecker=True
print_cmd=False
lock = threading.Lock()

def remove_cmd(cmd, target, with_param):
    poss = cmd.find(target)
    if poss < 0:
        return cmd
    
    pose = poss + len(target) + 1
    if with_param:
        while cmd[pose] == ' ':
            pose += 1
        flag = False
        while pose < len(cmd) and (flag or cmd[pose] != ' '):
            if cmd[pose] == '\"':
                flag = not(flag)
            pose += 1
    
    if pose < len(cmd):
        return cmd[0:poss] + cmd[pose:]
    else:
        return cmd[0:poss]


def kill_proc_tree(pid, including_parent=True):
    parent = psutil.Process(pid)
    children = parent.children(recursive=True)
    for child in children:
        child.kill()
    gone, still_alive = psutil.wait_procs(children, timeout=5)
    if including_parent:
        parent.kill()
        parent.wait(5)

class ProcessChecker:
    process = None
    cpuThreshold = 0.5

    def __init__(self, _process, _cpuThreshold):
        self.process = _process
        self.cpuThreshold = _cpuThreshold

    #プロセスの終了を待機し、終了コードを取得する
    #プロセスのCPU使用率が連続で十分低かったら、
    #プロセスが強制終了したとみなし、1を返す
    def wait_or_kill_if_dead(self, exename):
        ps = psutil.Process(self.process.pid)
        shellExeName = "cmd.exe" if os.name == 'nt' else "sh"
        if exename != shellExeName:
            try:
                while ps.name() == shellExeName:
                    for pchild in ps.children():
                        if exename in pchild.name():
                            ps = pchild
                    #プロセスが終了していなければ、Noneを返す
                    return_code = self.process.wait()
                    if return_code == None:
                        time.sleep(1)
                    else:
                        return ( int(return_code), False )
            except psutil.NoSuchProcess:
                time.sleep(1)
                return_code = self.process.wait()
                return ( int(return_code), False )
            except:
                raise

        cpu_usage_none_count = 0 #連続でCPU使用率が低い回数
        CPU_USAGE_NONE_MAX = 30 #プロセスが強制終了したとみなす、連続でCPU使用率が低い回数の閾値
        return_code = None #終了コード
        while return_code is None and cpu_usage_none_count < CPU_USAGE_NONE_MAX:
            try:
                cpu_usage = ps.cpu_percent(interval=1)
                if cpu_usage * float(psutil.cpu_count()) < self.cpuThreshold:
                    cpu_usage_none_count = cpu_usage_none_count + 1
                else:
                    cpu_usage_none_count = 0
            except:
                cpu_usage_none_count = cpu_usage_none_count + 1

            #プロセスが終了していなければ、Noneを返す
            return_code = self.process.poll()

        try:
            return_code = self.process.wait(10)
            return ( int(return_code), False )
        except:
            try:
                kill_proc_tree(self.process.pid)
            except:
                print("failed to kill encoder process")
            return ( 0, True )

class TestData:
    data_id = 0
    for_qsv = True
    for_nvenc = True
    for_vceenc = True
    for_rkmppenc = True
    command_line = ""
    inptut_file = ""
    output_prefix = ""
    comment = ""
    error_expected = False

    def __init__(self, _id, _qsv, _nvenc, _vceenc, _rkmppenc, _command_line, _input, _output_prefix, _comment, _error_expected):
        assert isinstance(_id, int)
        assert isinstance(_qsv, bool)
        assert isinstance(_nvenc, bool)
        assert isinstance(_vceenc, bool)
        assert isinstance(_rkmppenc, bool)
        assert isinstance(_command_line, str)
        assert isinstance(_input, str)
        assert isinstance(_output_prefix, str)
        assert isinstance(_comment, str)
        assert isinstance(_error_expected, bool)
        self.data_id = _id
        self.for_qsv = _qsv
        self.for_nvenc = _nvenc
        self.for_vceenc = _vceenc
        self.for_rkmppenc = _rkmppenc
        self.command_line = _command_line
        self.inptut_file = _input
        self.output_prefix = _output_prefix
        self.comment = _comment
        self.error_expected = _error_expected

class TestTable:
    list_test_data = []

    def __init__(self, input_xlsx):
        try:
            wb = openpyxl.load_workbook(filename=input_xlsx)
            ws = wb.active
        except:
            print("failed to load xlsx file : " + input_xlsx)
            print(traceback.format_exc())
            exit(1)

        self.read(ws)

    def cell_str(self, value):
        return "" if value is None else str(value).replace("○", "〇")

    def read(self, ws):
        y = 2 #openpyxlは1スタート
        data_id = 0
        no_data_row = 0
        THRESOLD_NO_DATA_ROW = 10 #行に連続でデータがなければ、データはもうないとみなす
        while no_data_row < THRESOLD_NO_DATA_ROW:
            if ws.cell(row = y, column = 5).value is None and ws.cell(row = y, column = 6).value is None:
                no_data_row = no_data_row + 1
            else:
                data_id = data_id + 1
                no_data_row = 0
                try:
                    for_qsv = self.cell_str(ws.cell(row = y, column = 1).value) == "〇"
                    for_nvenc = self.cell_str(ws.cell(row = y, column = 2).value) == "〇"
                    for_vceenc = self.cell_str(ws.cell(row = y, column = 3).value) == "〇"
                    for_rkmppenc = self.cell_str(ws.cell(row = y, column = 4).value) == "〇"
                    command_line = self.cell_str(ws.cell(row = y, column = 5).value)
                    inptut_file = self.cell_str(ws.cell(row = y, column = 6).value)
                    output_prefix = self.cell_str(ws.cell(row = y, column = 7).value)
                    comment = self.cell_str(ws.cell(row = y, column = 8).value)
                    error_expected = self.cell_str(ws.cell(row = y, column = 9).value) == "〇"
                except:
                    print("failed to parse xlsx file row " + str(y))
                    print(traceback.format_exc())
                    exit(1)

                if data_id >= test_start:
                    test_data = TestData(data_id, for_qsv, for_nvenc, for_vceenc, for_rkmppenc, command_line, inptut_file, output_prefix, comment, error_expected)
                    self.list_test_data.append(test_data)

            #次の行へ
            y = y + 1

        print("read " + str(len(self.list_test_data)) + " test datas.")

class ResultData:
    test_data = None
    ret_total = 0
    ret_enc_run = 0
    enc_killed = False
    ret_minfo_diff = 0
    ret_filesize = 0
    full_enc_cmd = ""

    def __init__(self, _test_data, _ret_enc_run, _enc_killed, _ret_minfo_diff, _ret_filesize, _full_enc_cmd):
        assert isinstance(_test_data, TestData)
        assert isinstance(_ret_enc_run, int)
        assert isinstance(_enc_killed, bool)
        assert isinstance(_ret_minfo_diff, int)
        assert isinstance(_full_enc_cmd, str)
        self.test_data = _test_data
        self.ret_enc_run = _ret_enc_run
        self.enc_killed = _enc_killed
        self.ret_minfo_diff = _ret_minfo_diff
        self.ret_filesize = _ret_filesize
        self.full_enc_cmd = _full_enc_cmd

        self.ret_total = 1
        if self.enc_killed == 0:
            if self.ret_enc_run != 0 and self.test_data.error_expected:
                self.ret_total = 0
            if self.ret_enc_run == 0 and self.ret_minfo_diff == 0 and self.ret_filesize == 0:
                self.ret_total = 0

    def write(self, output_xlsx):
        global lock
        try:
            lock.acquire()
            if os.path.exists(output_xlsx):
                wb = openpyxl.load_workbook(filename=output_xlsx)
                ws = wb.active
                ws.cell(row = 1, column =  1).value = "id"
                ws.cell(row = 1, column =  2).value = "ret total"
                ws.cell(row = 1, column =  3).value = "ret enc run"
                ws.cell(row = 1, column =  4).value = "ret enc killed"
                ws.cell(row = 1, column =  5).value = "ret media info"
                ws.cell(row = 1, column =  6).value = "ret file size"
                ws.cell(row = 1, column =  7).value = "error expected"
                ws.cell(row = 1, column =  8).value = "command line"
                ws.cell(row = 1, column =  9).value = "input file"
                ws.cell(row = 1, column = 10).value = "output prefix"
                ws.cell(row = 1, column = 11).value = "comment"
                ws.cell(row = 1, column = 12).value = "for_qsv"
                ws.cell(row = 1, column = 13).value = "for_nvenc"
                ws.cell(row = 1, column = 14).value = "for_vceenc"
                ws.cell(row = 1, column = 15).value = "for_rkmppenc"
                ws.cell(row = 1, column = 16).value = "full_enc_cmd"
                ws.column_dimensions[openpyxl.utils.get_column_letter( 1)].width = 5
                ws.column_dimensions[openpyxl.utils.get_column_letter( 2)].width = 3
                ws.column_dimensions[openpyxl.utils.get_column_letter( 3)].width = 3
                ws.column_dimensions[openpyxl.utils.get_column_letter( 4)].width = 3
                ws.column_dimensions[openpyxl.utils.get_column_letter( 5)].width = 3
                ws.column_dimensions[openpyxl.utils.get_column_letter( 6)].width = 3
                ws.column_dimensions[openpyxl.utils.get_column_letter( 7)].width = 3
                ws.column_dimensions[openpyxl.utils.get_column_letter( 8)].width = 60
                ws.column_dimensions[openpyxl.utils.get_column_letter( 9)].width = 50
                ws.column_dimensions[openpyxl.utils.get_column_letter(11)].width = 40
                ws.column_dimensions[openpyxl.utils.get_column_letter(12)].width = 3
                ws.column_dimensions[openpyxl.utils.get_column_letter(13)].width = 3
                ws.column_dimensions[openpyxl.utils.get_column_letter(14)].width = 3
                ws.column_dimensions[openpyxl.utils.get_column_letter(15)].width = 3
                ws.column_dimensions[openpyxl.utils.get_column_letter(16)].width = 3
            else:
                wb = openpyxl.Workbook() #新しいworkbook
                ws = wb.active
            try:
                y = 2 #openpyxlは1スタート
                no_data_row = 0
                THRESOLD_NO_DATA_ROW = 10
                while ws.cell(row = y, column = 1).value is not None:
                    y = y + 1

                ws.cell(row = y, column =  1).value = str(self.test_data.data_id)
                ws.cell(row = y, column =  2).value = ("×" if self.ret_total != 0 else "")
                ws.cell(row = y, column =  3).value = ("×" if self.ret_enc_run != 0 else "")
                ws.cell(row = y, column =  4).value = ("×" if self.enc_killed != 0 else "")
                ws.cell(row = y, column =  5).value = ("×" if self.ret_minfo_diff != 0 else "")
                ws.cell(row = y, column =  6).value = ("×" if self.ret_filesize != 0 else "")
                ws.cell(row = y, column =  7).value = ("×" if self.test_data.error_expected else "")
                ws.cell(row = y, column =  8).value = self.test_data.command_line
                ws.cell(row = y, column =  9).value = self.test_data.inptut_file
                ws.cell(row = y, column = 10).value = self.test_data.output_prefix
                ws.cell(row = y, column = 11).value = self.test_data.comment
                ws.cell(row = y, column = 12).value = ("〇" if self.test_data.for_qsv else "")
                ws.cell(row = y, column = 13).value = ("〇" if self.test_data.for_nvenc else "")
                ws.cell(row = y, column = 14).value = ("〇" if self.test_data.for_vceenc else "")
                ws.cell(row = y, column = 15).value = ("〇" if self.test_data.for_rkmppenc else "")
                ws.cell(row = y, column = 16).value = ("〇" if self.full_enc_cmd else "")
                ws.cell(row = y, column =  8).number_format = openpyxl.styles.numbers.FORMAT_TEXT
                ws.cell(row = y, column =  9).number_format = openpyxl.styles.numbers.FORMAT_TEXT
                ws.cell(row = y, column = 10).number_format = openpyxl.styles.numbers.FORMAT_TEXT
                ws.cell(row = y, column = 11).number_format = openpyxl.styles.numbers.FORMAT_TEXT
                try:
                    wb.save(output_xlsx)
                except:
                    print("failed to update xlsx file : " + output_xlsx)
                    print(traceback.format_exc())
            except:
                print("failed set result to xlsx file : " + output_xlsx)
                print(traceback.format_exc())
        except:
            print("failed to open xlsx file : " + output_xlsx)
            print(traceback.format_exc())
        finally:
            lock.release()

class ResultTable:
    list_result_data = []

    def __init__(self, input_xlsx):
        try:
            wb = openpyxl.load_workbook(filename=input_xlsx)
            ws = wb.active
        except:
            print("failed to load xlsx file : " + input_xlsx)
            print(traceback.format_exc())
            exit(1)

        self.read(ws)

    def cell_str(self, value):
        return "" if value is None else str(value).replace("○", "〇")

    def read(self, ws):
        y = 2 #openpyxlは1スタート
        data_id = 0
        no_data_row = 0
        THRESOLD_NO_DATA_ROW = 10
        while no_data_row < THRESOLD_NO_DATA_ROW:
            if ws.cell(row = y, column = 1).value is None:
                no_data_row = no_data_row + 1
            else:
                data_id = data_id + 1
                no_data_row = 0
                try:
                    data_id = int(ws.cell(row = y, column =  1).value)
                    ret_total = 1 if self.cell_str(ws.cell(row = y, column = 2).value) == "×" else 0
                    ret_enc_run = 1 if self.cell_str(ws.cell(row = y, column = 3).value) == "×" else 0
                    enc_killed = 1 if self.cell_str(ws.cell(row = y, column = 4).value) == "×" else 0
                    ret_minfo_diff = 1 if self.cell_str(ws.cell(row = y, column = 5).value) == "×" else 0
                    ret_filesize = 1 if self.cell_str(ws.cell(row = y, column = 6).value) == "×" else 0
                    error_expected = self.cell_str(ws.cell(row = y, column = 7).value) == "×"
                    command_line = self.cell_str(ws.cell(row = y, column = 8).value)
                    inptut_file = self.cell_str(ws.cell(row = y, column = 9).value)
                    output_prefix = self.cell_str(ws.cell(row = y, column = 10).value)
                    comment = self.cell_str(ws.cell(row = y, column = 11).value)
                    for_qsv = self.cell_str(ws.cell(row = y, column = 12).value) == "〇"
                    for_nvenc = self.cell_str(ws.cell(row = y, column = 13).value) == "〇"
                    for_vceenc = self.cell_str(ws.cell(row = y, column = 14).value) == "〇"
                    for_rkmppenc = self.cell_str(ws.cell(row = y, column = 15).value) == "〇"
                    full_enc_cmd = self.cell_str(ws.cell(row = y, column = 16).value)
                except:
                    print("failed to parse xlsx file row " + str(y))
                    print(traceback.format_exc())
                    exit(1)

                test_data = TestData(data_id, for_qsv, for_nvenc, for_vceenc, for_rkmppenc, command_line, inptut_file, output_prefix, comment, error_expected)
                result_data = ResultData(test_data, ret_enc_run, enc_killed, ret_minfo_diff, ret_filesize, full_enc_cmd)
                self.list_result_data.append(test_data)
                y = y + 1

class HWEncTest:
    encoder_path = ""
    encoder_name = ""
    encoder_log_path = ""
    mediainfo_compare_dir = ""
    def __init__(self, _encoder_path, _encoder_name, _encoder_log_path, _mediainfo_compare_dir):
        assert isinstance(_encoder_path, str)
        assert isinstance(_encoder_name, str)
        assert isinstance(_encoder_log_path, str)
        assert isinstance(_mediainfo_compare_dir, str)
        self.encoder_path = _encoder_path
        self.encoder_name = _encoder_name
        self.encoder_log_path = _encoder_log_path
        self.mediainfo_compare_dir = _mediainfo_compare_dir

    def check_if_run_required(self, test_data):
        assert isinstance(test_data, TestData)
        if encoder_name == "qsvencc" and test_data.for_qsv:
            return True
        if encoder_name == "nvencc" and test_data.for_nvenc:
            return True
        if encoder_name == "vceencc" and test_data.for_vceenc:
            return True
        if encoder_name == "rkmppenc" and test_data.for_rkmppenc:
            return True
        return False

    def replace_cmd(self, test_data, cmd):
        assert isinstance(test_data, TestData)
        cmd_ffmpeg = ("$(ExePath)" in test_data.command_line) and ("$(FFmpegPath)" in test_data.command_line) and (test_data.command_line.find("$(ExePath)") < test_data.command_line.find("$(FFmpegPath)"))
        cmd = cmd.replace("$(OutDir)", outputdir)
        cmd = cmd.replace("$(InputFile)", test_data.inptut_file)
        cmd = cmd.replace("$(OutputFile)", self.output_file_path(test_data))
        cmd = cmd.replace("$(LogFile)", self.log_file_path(test_data))
        cmd = cmd.replace("$(ExePath)", self.encoder_path)
        cmd = cmd.replace("$(FFmpegPath)", ffmpeg_path)
        if encoder_name == "nvencc":
            cmd = cmd.replace("--d3d11", "")
            cmd = cmd.replace("--d3d9", "")
            cmd = cmd.replace("--disable-d3d", "")
            cmd = cmd.replace("-u 7", "--preset performance")
            cmd = cmd.replace("--avqsv", "--avhw")
            if not cmd_ffmpeg:
                cmd = cmd + " --gpu-select cores=0.0,gen=0.0,gpu=0.5"
        elif encoder_name == "vceencc":
            cmd = cmd.replace("-u 7", "")
            if not '--vpp-afs' in cmd and not '--vpp-yadif' in cmd and not '--vpp-nnedi' in cmd:
                cmd = cmd.replace("--tff", "")
                cmd = cmd.replace("--bff", "")
            cmd = cmd.replace("--d3d11", "")
            cmd = cmd.replace("--d3d9", "")
            cmd = cmd.replace("--disable-d3d", "")
            cmd = cmd.replace("--vpp-deinterlace normal", "")
            cmd = cmd.replace("--vpp-deinterlace bob", "")
            cmd = cmd.replace("--avqsv", "--avhw")
            cmd = cmd.replace("--profile main10", "--output-depth 10")
            cmd = cmd.replace("--avhw", "")
            if not cmd_ffmpeg:
                cmd += " -d 0 "
            #cmd = remove_cmd(cmd, "--trim", True)
        elif encoder_name == "rkmppenc":
            cmd = cmd.replace("-u 7", "")
            if not '--vpp-afs' in cmd and not '--vpp-yadif' in cmd and not '--vpp-nnedi' in cmd:
                cmd = cmd.replace("--tff", "")
                cmd = cmd.replace("--bff", "")
            cmd = cmd.replace("--d3d11", "")
            cmd = cmd.replace("--d3d9", "")
            cmd = cmd.replace("--disable-d3d", "")
            cmd = cmd.replace("--vpp-deinterlace normal", "--vpp-deinterlace normal_i5")
            cmd = cmd.replace("--vpp-deinterlace bob", "--vpp-deinterlace bob_i5")
            cmd = cmd.replace("--avqsv", "--avhw")
        return cmd

    def output_file_path(self, test_data):
        assert isinstance(test_data, TestData)
        output_file = test_data.output_prefix
        if output_file != "-" and test_data.inptut_file != "-":
            output_file = os.path.join(outputdir, "{0:04d}".format(test_data.data_id) + "_" + test_data.inptut_file.replace(':', '_') + test_data.output_prefix)
        return output_file

    def log_file_path(self, test_data):
        assert isinstance(test_data, TestData)
        return self.output_file_path(test_data) + encoder_log_prefix

    def generate_enc_cmd(self, test_data):
        assert isinstance(test_data, TestData)
        add_exepath = ("$(ExePath)" not in test_data.command_line)
        add_input = ("$(InputFile)" not in test_data.command_line) and len(test_data.inptut_file) > 0
        add_output = ("$(OutputFile)" not in test_data.command_line) and len(test_data.output_prefix) > 0
        add_logout = ("$(LogFile)" not in test_data.command_line)
        cmd_ffmpeg = ("$(ExePath)" in test_data.command_line) and ("$(FFmpegPath)" in test_data.command_line) and (test_data.command_line.find("$(ExePath)") < test_data.command_line.find("$(FFmpegPath)"))

        cmd = ""
        if add_exepath:
            cmd = "\"" + self.encoder_path + "\" "

        cmd = cmd + self.replace_cmd(test_data, test_data.command_line)
        if add_logout:
            cmd = cmd + " --log \"" + self.log_file_path(test_data) + "\""

        if add_output:
            cmd = cmd + " -o \"" + self.output_file_path(test_data) + "\""

        if add_input:
            cmd = cmd + " -i \"" + test_data.inptut_file + "\""
            if encoder_name == "vceencc" and "君の名は。　ULTRA_HD_t00.mkv" == test_data.inptut_file: # 特例対応 (avhwだとエラーになるので)
                cmd = cmd.replace("--avhw", "--avsw")
                cmd = cmd + " --avsww"
                
        if not cmd_ffmpeg:
            if encoder_name != "rkmppenc":
                cmd = cmd + " --thread-affinity main=ecore,decoder=ecore,output=ecore,audio=ecore,perfmonitor=ecore"
            cmd = cmd + " --no-mp4opt"

        # cmd = cmd + " --log-level debug"

        if os.name == "posix":
            cmd = cmd.replace(";", "\\;") #エスケープが必要
        return cmd

    def run_encoder(self, test_data):
        assert isinstance(test_data, TestData)

        cmd = self.generate_enc_cmd(test_data)
        if print_cmd:
            print(cmd)
        killed = False

        try:
            p = subprocess.Popen(cmd, stderr = subprocess.DEVNULL, shell=True)
            if UseProcessChecker:
                cpuThreshold = 0.5
                if self.encoder_name == "qsvencc":
                    cpuThreshold = 0.5
                elif self.encoder_name == "nvencc":
                    cpuThreshold = 0.01
                elif self.encoder_name == "vceencc":
                    cpuThreshold = 0.01
                proc_check = ProcessChecker(p, cpuThreshold)
                ret, killed = proc_check.wait_or_kill_if_dead(os.path.basename(self.encoder_path))
            else:
                ret = p.wait()
                killed = False
        except:
            print("failed to run encoder\n");
            print(traceback.format_exc())
            ret = 1

        return ( ret, killed )

    def run_mediainfo(self, test_data):
        assert isinstance(test_data, TestData)
        with open(self.output_file_path(test_data) + mediainfo_check_log_appendix, 'w') as outfile:
            cmd = "\"" + mediainfo_path + "\"" \
                + " --Output=file://" + mediainfo_template + " " \
                + "\"" + self.output_file_path(test_data) + "\""
            try:
                p = subprocess.run(cmd, stdout=outfile, shell=True)
                ret = p.returncode
            except:
                ret = 1
            return ret

    def compare_mediainfo(self, test_data):
        assert isinstance(test_data, TestData)
        if self.mediainfo_compare_dir is None or len(self.mediainfo_compare_dir) == 0:
            return 0

        minfo_log_current = self.output_file_path(test_data) + mediainfo_check_log_appendix
        minfo_log_compare = os.path.join(self.mediainfo_compare_dir, \
            "{0:04d}".format(test_data.data_id) + "_" \
            + ("" if test_data.inptut_file == "-" else test_data.inptut_file) \
            + test_data.output_prefix + mediainfo_check_log_appendix)

        try:
            #mediainfoの出力はUTF-8
            document_compare = codecs.open(minfo_log_compare, 'r', 'utf-8')
        except:
            print("failed to open file: " + minfo_log_compare);
            print(traceback.format_exc())
            return 1

        try:
            #mediainfoの出力はUTF-8
            document_current = codecs.open(minfo_log_current, 'r', 'utf-8')
        except:
            print("failed to open file: " + minfo_log_current);
            print(traceback.format_exc())
            return 1

        #改行コードは一度LFのみに変換して単純化、splitlinesはTrueを指定して改行コードを残す
        text_compare = document_compare.read().replace('\r\n', '\n').splitlines(True)
        text_current = document_current.read().replace('\r\n', '\n').splitlines(True)
        mediainfo_log_diff = difflib.unified_diff(text_compare, text_current, minfo_log_compare, minfo_log_current, lineterm='\n')

        diff_file_path = self.output_file_path(test_data) + mediainfo_check_diff_appendix

        try:
            diff_file = codecs.open(diff_file_path, 'w', 'utf-8')
            for diff_line in mediainfo_log_diff:
                #出力時にCRLFに戻す
                diff_file.write(diff_line.replace('\n', '\r\n'))
        except:
            print("failed to open file: " + diff_file_path);
            print(traceback.format_exc())
            return 1

        ratio = difflib.SequenceMatcher(None, text_compare, text_current).ratio()
        return 0 if ratio == 1.0 else 1

    def compare_filesize(self, test_data):
        assert isinstance(test_data, TestData)
        if self.mediainfo_compare_dir is None or len(self.mediainfo_compare_dir) == 0:
            return 0

        out_file_current = self.output_file_path(test_data)
        out_file_compare = os.path.join(self.mediainfo_compare_dir, \
            "{0:04d}".format(test_data.data_id) + "_" \
            + ("" if test_data.inptut_file == "-" else test_data.inptut_file) \
            + test_data.output_prefix)

        try:
            size_current = os.path.getsize(out_file_current)
        except:
            print("failed to get file size of output file: " + out_file_current);
            print(traceback.format_exc())
            return 1

        try:
            size_compare = os.path.getsize(out_file_compare)
        except:
            print("failed to get file size of output file: " + out_file_compare);
            print(traceback.format_exc())
            return 1

        if size_current == 0:
            return 1

        if size_compare == 0:
            return 1

        ratio = abs(1.0 - float(size_current) / float(size_compare))
        return 0 if ratio < filesize_threshold else 1

    def run_test(self, test_data):
        assert isinstance(test_data, TestData)
        if not self.check_if_run_required(test_data):
            return True

        #print("-------------------------------------------------------------------------------")
        #print("start test #" + str(test_data.data_id))

        ret_enc_run, enc_killed = self.run_encoder(test_data)
        ret_minfo_run = 0
        ret_minfo_diff = 0
        ret_file_size = 0
        
        if not "longpathtest" in test_data.inptut_file:
            if ret_enc_run == 0:
                if not os.path.exists(self.output_file_path(test_data)):
                    ret_enc_run = 1

            if ret_enc_run == 0:
                ret_minfo_run = self.run_mediainfo(test_data)

                if ret_minfo_run == 0:
                    ret_minfo_diff = self.compare_mediainfo(test_data)
                    ret_file_size = self.compare_filesize(test_data)

            if "option_check" not in test_data.comment:
                try:
                    fp_enc_log = open(self.log_file_path(test_data), "rb")
                    log_lines = fp_enc_log.read()
                    fp_enc_log.close()

                    fp_enc_log = open(self.encoder_log_path, "a")
                    fp_enc_log.writelines("-------------------------------------------------------------------------------\n")
                    fp_enc_log.writelines("start test #" + str(test_data.data_id) + "\n")
                    fp_enc_log.close()

                    fp_enc_log = open(self.encoder_log_path, "ab")
                    fp_enc_log.write(log_lines)
                    fp_enc_log.close()
                except:
                    print("error opening " + encoder_name + " log file.\n")
                    print(traceback.format_exc())

        result_data = ResultData(test_data, ret_enc_run, enc_killed, ret_minfo_diff, ret_file_size, self.generate_enc_cmd(test_data))
        result_data.write(os.path.join(outputdir, output_xlsx_filename))
        print("test #" + str(test_data.data_id) + " result: " + ("〇" if (result_data.ret_total == 0) else "×"))

        return True if (result_data.ret_total == 0) else False


if __name__ == '__main__':
    sleep_after_run = False
    print(sys.version_info)

    mediainfo_compare_dir = ""
    input_xlsx = ""
    process = 1
    
    if os.name == 'nt':
        x264_path = x264_path_win
        avs2pipemod_path = avs2pipemod_path_win
        qsvencc_path = qsvencc_path_win
        nvencc_path = nvencc_path_win
        vceencc_path = vceencc_path_win
        ffmpeg_path = ffmpeg_path_win
        mediainfo_path = mediainfo_path_win
        computer_name = os.environ.get("COMPUTERNAME")
    else:
        computer_name = subprocess.check_output("hostname", shell=True).decode('utf-8').strip().replace('-', '_')
        print(computer_name)

    iarg = 0
    while iarg < len(sys.argv):
        if sys.argv[iarg] == "-s":
            sleep_after_run = True
            print("sleep_after_run")
        elif sys.argv[iarg] == "-o":
            iarg=iarg+1
            outputdir = sys.argv[iarg]
        elif sys.argv[iarg] == "-p":
            iarg=iarg+1
            process = int(sys.argv[iarg])
        elif sys.argv[iarg] == "-q":
            encoder_path = qsvencc_path
        elif sys.argv[iarg] == "-qp":
            iarg=iarg+1
            qsvencc_path=sys.argv[iarg]
            encoder_path = qsvencc_path
        elif sys.argv[iarg] == "-compare-dir":
            iarg=iarg+1
            mediainfo_compare_dir = sys.argv[iarg]
        elif sys.argv[iarg] == "-n":
            encoder_path = nvencc_path
        elif sys.argv[iarg] == "-np":
            iarg=iarg+1
            nvencc_path=sys.argv[iarg]
            encoder_path = nvencc_path
        elif sys.argv[iarg] == "-v":
            encoder_path = vceencc_path
        elif sys.argv[iarg] == "-vp":
            iarg=iarg+1
            vceencc_path=sys.argv[iarg]
            encoder_path = vceencc_path
        elif sys.argv[iarg] == "-r":
            encoder_path = rkmppenc_path
        elif sys.argv[iarg] == "-rp":
            iarg=iarg+1
            rkmppenc_path=sys.argv[iarg]
            encoder_path = rkmppenc_path
        elif sys.argv[iarg] == "-ts":
            iarg=iarg+1
            test_start = int(sys.argv[iarg])
        elif sys.argv[iarg] == "-t":
            iarg=iarg+1
            test_target = int(sys.argv[iarg])
        elif sys.argv[iarg] == "-nc":
            UseProcessChecker=False
        elif sys.argv[iarg] == "-x":
            iarg=iarg+1
            input_xlsx=sys.argv[iarg]
        elif sys.argv[iarg] == "-print-cmd":
            print_cmd = True
        iarg=iarg+1

    if encoder_path == qsvencc_path:
        encoder_name = "qsvencc"
    elif encoder_path == nvencc_path:
        encoder_name = "nvencc"
    elif encoder_path == vceencc_path:
        encoder_name = "vceencc"
    elif encoder_path == rkmppenc_path:
        encoder_name = "rkmppenc"
    else:
        print("unknown encoder path set")
        exit(1)

    py_path, py_ext = os.path.splitext(os.path.basename(__file__))
    if len(input_xlsx) == 0:
        input_xlsx = py_path + ".xlsx"
    outputdir = os.path.join(outputdir, "output_" + computer_name + "_" + encoder_name + "_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
    logpath = encoder_name + "_test_result_" + computer_name + ".csv"
    encoder_log_path = encoder_name + "_test_" + computer_name + ".txt"
    print("encoder selected:" + encoder_name)

    if os.path.isdir(outputdir):
        shutil.rmtree(outputdir)
    os.mkdir(outputdir)
    if os.path.exists(logpath):
        os.remove(logpath)
    if os.path.exists(encoder_log_path):
        os.remove(encoder_log_path)

    test_table = TestTable(input_xlsx)

    test = HWEncTest(encoder_path, encoder_name, encoder_log_path, mediainfo_compare_dir)

    result = joblib.Parallel(n_jobs=process, backend='threading')( \
        [joblib.delayed(test.run_test)(test_data) for test_data in \
            [test_data for test_data in test_table.list_test_data if (test_target == 0 or test_data.data_id == test_target)] \
        ])

    if sleep_after_run:
        ctypes.windll.PowrProf.SetSuspendState(0, 1, 0)
